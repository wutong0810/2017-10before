# -*- coding: utf-8 -*-
"""
Created on Sun May 07 22:11:22 2017

@author: Administrator
"""

import os
path1=r'D:\KuGou'
path2=unicode(path1,"utf8")
filename_total=[]
for dirpath, dirnames, filenames in os.walk(path2):
    for filename in filenames:
        
        if os.path.splitext(filename)[1]=='.xlsx':
            filename2=os.path.join(dirpath,filename)
            print filename2
            filename_total.append(filename2)
filename_total     


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys
 
def xlsx2csv(filename,csv_filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            print sheet
#     # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
#         csv_filename = '{xlsx}_{sheet}.csv'.format(
#         xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
#         sheet=sheet.replace(' ', '_'))
            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)
            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    print cell.value
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
        csv_file.close()
    except Exception as e:
        print(e)
        
        

        
        
        
        
        
        
        
        
        
        
import os
path1=r'F:\萧山卡口流量0410-0416'
path2=unicode(path1,"utf8")
filename_total=[]
for dirpath, dirnames, filenames in os.walk(path2):
    for filename in filenames:
        
        if os.path.splitext(filename)[1]=='.xls':
            filename2=os.path.join(dirpath,filename)
            print filename2
            filename_total.append(filename2)              
csv_filename=r'd:1.csv'        
for filename in filename_total:        
    xlsx_file_reader = load_workbook(filename=filename)
    for sheet in xlsx_file_reader.get_sheet_names():
#        print sheet
    #     # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
    #         csv_filename = '{xlsx}_{sheet}.csv'.format(
    #         xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
    #         sheet=sheet.replace(' ', '_'))
        csv_file = file(csv_filename, 'wb')
        csv_file_writer = csv.writer(csv_file)
        sheet_ranges = xlsx_file_reader[sheet]
        for row in sheet_ranges.rows:
            row_container = []
            for cell in row:
#                print cell.value
                if type(cell.value) == unicode:
                    row_container.append(cell.value.encode('utf-8'))
                else:
                    row_container.append(str(cell.value))
            csv_file_writer.writerow(row_container)
csv_file.close()        
        
        
import xlrd
import xlwt
import sys
from datetime import date,datetime        
import os
path1=r'F:\萧山卡口流量0410-0416'
path2=unicode(path1,"utf8")
filename_total=[]
for dirpath, dirnames, filenames in os.walk(path2):
    for filename in filenames:
        if os.path.splitext(filename)[1]=='.xls':
            filename2=os.path.join(dirpath,filename)
            print filename2
            filename_total.append(filename2)              
csv_filename=r'd:/1.csv' 
csv_file = file(csv_filename, 'wb')
csv_file_writer = csv.writer(csv_file)       
for filename in filename_total:           
    try:    
        workbook = xlrd.open_workbook(filename,encoding_override="utf-8")   
        sheet2 = workbook.sheet_by_index(0)   
        for row in xrange(0, sheet2.nrows):
            rows = sheet2.row_values(row)
            row_container = []
            for cell in rows:
                if type(cell) == unicode:
                    row_container.append(cell.encode('utf8'))
                else:
                    row_container.append(str(cell))
            csv_file_writer.writerow(row_container)
    except Exception as e:
            print(e)
csv_file.close()       
    
    
    
    
    
      if type(u'') == type(cell): 
        return "\"%s\"" % cell.encode('utf8')
      else:
        return "\"%s\"" % str(cell) 
   
    print ','.join([_tostr(cell) for cell in rows ])        
        
        
        
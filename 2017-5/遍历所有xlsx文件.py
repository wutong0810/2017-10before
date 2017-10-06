# -*- coding: utf-8 -*-
"""
Created on Mon May 08 11:24:03 2017

@author: Administrator
"""
#获取 文件夹下所有的数据
import os
path1=r'D:\KuGou'
path2=unicode(path1,"utf8")
filename_total=[]
for dirpath, dirnames, filenames in os.walk(path2):
    for filename in filenames:
        
        if os.path.splitext(filename)[1]=='.xlsx':
            filename2=os.path.join(dirpath,filename)
            print filename2
            filename_total.append(filename2)

#读写所有数据
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys
 
def xlsx2csv(filename,csv_filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            print sheet
#     # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
#         csv_filename = '{xlsx}_{sheet}.csv'.format(
#         xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
#         sheet=sheet.replace(' ', '_'))
            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)
            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    print cell.value
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
        csv_file.close()
    except Exception as e:
        print(e)